from flask import Flask, request, jsonify
from flask_cors import CORS
from dotenv import load_dotenv
import os
import json
import time
import threading
import requests
from groq import Groq
import openpyxl
from openpyxl import load_workbook
from pathlib import Path

load_dotenv()
app = Flask(__name__)
CORS(app)

# ── Constants ──────────────────────────────────────────────────────────────────
FILLER_WORDS = ["um", "uh", "er", "hmm", "like", "you know", "basically",
                "literally", "actually", "so", "right", "okay so", "i mean"]

CEFR_RANGES = {
    "A1": (0, 10), "A2": (11, 20), "B1": (21, 30),
    "B2": (31, 40), "C1": (41, 50), "C2": (51, 60)
}

API_KEYS = {
    "Groq (Llama 3.3 70B)":  os.getenv("GROQ_API_KEY"),
    "Gemini 2.5 Flash Lite":       os.getenv("GEMINI_API_KEY"),
    "Ollama (Gemma4 31B)":       os.getenv("OLLAMA_API_KEY"),
    
}

# ── Helpers ────────────────────────────────────────────────────────────────────
def get_cefr(total):
    for level, (lo, hi) in CEFR_RANGES.items():
        if lo <= total <= hi:
            return level
    return "A1"

def extract_features(transcript, duration):
    words = transcript.split()
    word_count = len(words)
    wpm = round((word_count / duration) * 60) if duration > 0 else 0
    filler_count = sum(transcript.lower().count(f) for f in FILLER_WORDS)
    return {
        "word_count": word_count,
        "duration_seconds": round(duration, 1),
        "wpm": wpm,
        "filler_count": filler_count,
    }

def build_prompt(transcript, features):
    return f"""You are an expert CEFR communication analyst and English language evaluator.
Analyze this speech sample and score the speaker strictly on each parameter from 0 to 5.

=== SPEECH DATA ===
TRANSCRIPT:
\"{transcript}\"

SPEECH METRICS:
- Duration: {features['duration_seconds']}s
- Word Count: {features['word_count']}
- Speaking Rate: {features['wpm']} WPM (ideal: 120-150)
- Filler Word Count: {features['filler_count']}

=== SCORING RUBRIC (0-5 each) ===
0 = Very Poor | 1 = Poor | 2 = Below Average | 3 = Average | 4 = Good | 5 = Excellent

Parameters:
1. Vowel Sounds - Accuracy and clarity of vowel pronunciation
2. Consonant Sounds - Accuracy and clarity of consonant pronunciation
3. Word Stress - Correct stress patterns on syllables/words
4. Intonation - Natural rising/falling patterns, expressiveness
5. Sentence Construction - Grammar, syntax, complexity
6. Comprehensibility (Accent) - How easily understood despite accent
7. Clarity of Speech - Articulation, diction, enunciation
8. Projection of Voice - Volume, resonance, audibility
9. Avoided Fillers/Foghorns - Absence of um/uh/like/you know (5=zero fillers)
10. Confidence - Assertiveness, conviction in delivery
11. Rate of Speech - Appropriate pace (120-150 WPM is ideal)
12. Thought Flow - Logical organization, coherence, transitions

=== OUTPUT FORMAT ===
Return ONLY valid JSON, no explanation, no markdown:
{{
  "scores": {{
    "Vowel Sounds": <0-5>,
    "Consonant Sounds": <0-5>,
    "Word Stress": <0-5>,
    "Intonation": <0-5>,
    "Sentence Construction": <0-5>,
    "Comprehensibility (Accent)": <0-5>,
    "Clarity of Speech": <0-5>,
    "Projection of Voice": <0-5>,
    "Avoided Fillers/Foghorns": <0-5>,
    "Confidence": <0-5>,
    "Rate of Speech": <0-5>,
    "Thought Flow": <0-5>
  }},
  "total": <sum of all scores 0-60>,
  "cefr": "<A1|A2|B1|B2|C1|C2>",
  "summary": "<2 sentence analyst summary with areas of improvement>"
}}"""

# ── Transcribe using Groq Whisper API ─────────────────────────────────────────
def transcribe_audio(audio_file):
    groq_key = API_KEYS["Groq (Llama 3.3 70B)"]
    if not groq_key:
        raise Exception("Groq API key not found in .env")

    client = Groq(api_key=groq_key)
    start = time.time()

    transcription = client.audio.transcriptions.create(
        file=(audio_file.filename, audio_file.read(), audio_file.content_type),
        model="whisper-large-v3-turbo",
        response_format="verbose_json",
    )

    duration = time.time() - start
    transcript = transcription.text.strip()
    audio_duration = getattr(transcription, "duration", duration)

    return transcript, audio_duration

# ── LLM Callers ────────────────────────────────────────────────────────────────
def call_groq(prompt, api_key):
    r = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": "llama-3.3-70b-versatile",
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.2, "max_tokens": 800,
            "response_format": {"type": "json_object"}
        }, timeout=120
    )
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]

def call_gemini(prompt, api_key):
    r = requests.post(
        f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-lite:generateContent?key={api_key}",
        json={
            "contents": [{"parts": [{"text": prompt}]}],
            "generationConfig": {
                "temperature": 0.2,
                "maxOutputTokens": 2000,
            }
        }, timeout=120
    )
    r.raise_for_status()
    raw = r.json()["candidates"][0]["content"]["parts"][0]["text"]
    print("GEMINI RAW:", raw)  # shows in Flask terminal
    raw = raw.strip()
    if "```" in raw:
        raw = raw.replace("```json", "").replace("```", "").strip()
    start = raw.find("{")
    end = raw.rfind("}") + 1
    if start != -1 and end > start:
        raw = raw[start:end]
    return raw

def call_ollama(prompt, api_key):
    from ollama import Client
    client = Client(
        host="https://ollama.com",
        headers={"Authorization": f"Bearer {api_key}"}
    )
    response = client.chat(
        model="gemma4:31b",
        messages=[{"role": "user", "content": prompt}],
        stream=False,
        format="json"
    )
    return response["message"]["content"]
    print("OLLAMA RAW:", repr(content))  # debug
    return content

def call_cerebras(prompt, api_key):
    r = requests.post(
        "https://api.cerebras.ai/v1/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": "llama3.3-70b",
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.2, "max_tokens": 800,
            "response_format": {"type": "json_object"}
        }, timeout=30
    )
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]

def call_deepseek(prompt, api_key):
    r = requests.post(
        "https://api.deepseek.com/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": "deepseek-chat",
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.2, "max_tokens": 800,
            "response_format": {"type": "json_object"}
        }, timeout=30
    )
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]

def call_mistral(prompt, api_key):
    r = requests.post(
        "https://api.mistral.ai/v1/chat/completions",
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        json={
            "model": "mistral-small-latest",
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.2, "max_tokens": 800,
            "response_format": {"type": "json_object"}
        }, timeout=30
    )
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]

CALLERS = {
    "Groq (Llama 3.3 70B)":  call_groq,
    "Gemini 2.5 Flash Lite": call_gemini,
    "Ollama (Gemma4 31B)":       call_ollama,
     
}

def score_with_llm(name, prompt, api_key, results_dict):
    start = time.time()
    try:
        raw = CALLERS[name](prompt, api_key)
        data = json.loads(raw)
        data["latency"] = round(time.time() - start, 2)
        data["status"] = "done"
        if "scores" in data:
            data["total"] = sum(data["scores"].values())
            data["cefr"] = get_cefr(data["total"])
    except Exception as e:
        data = {"status": "error", "error": str(e), "latency": round(time.time() - start, 2)}
    results_dict[name] = data

def build_improvement_prompt(transcript, weak_params, all_scores):
    params_text = "\n".join(
        f"- {p}: scored {s}/5" for p, s in weak_params.items()
    )
    return f"""You are an expert CEFR communication coach. A student has been evaluated on their spoken English and the following areas need improvement.

TRANSCRIPT:
"{transcript}"

WEAK AREAS (scored 2 or below out of 5):
{params_text}

For EACH weak area listed above, give the student:
1. A brief explanation of why it matters
2. One specific, actionable tip they can practice today
3. A simple exercise or technique to improve it

Keep each tip practical and encouraging, not harsh. Maximum 2-3 sentences per area.

Return ONLY valid JSON, no markdown, no explanation:
{{
  "improvements": [
    {{
      "parameter": "<parameter name>",
      "why_it_matters": "<1 sentence>",
      "tip": "<actionable tip>",
      "exercise": "<simple practice exercise>"
    }}
  ],
  "overall_encouragement": "<2 sentence motivating message about their progress and potential>"
}}"""

def save_to_excel(username, llm_results, features):
    EXCEL_PATH = Path("E:/cefr_comparator/backend/results.xlsx")
    PARAMETERS = [
        "Vowel Sounds", "Consonant Sounds", "Word Stress", "Intonation",
        "Sentence Construction", "Comprehensibility (Accent)", "Clarity of Speech",
        "Projection of Voice", "Avoided Fillers/Foghorns", "Confidence",
        "Rate of Speech", "Thought Flow"
    ]

    # Calculate averages across all LLMs
    done = {k: v for k, v in llm_results.items() if v.get("status") == "done"}
    if not done:
        return

    avg_scores = {}
    for param in PARAMETERS:
        avg_scores[param] = round(
            sum(v["scores"].get(param, 0) for v in done.values()) / len(done), 1
        )

    total = round(sum(avg_scores.values()), 1)
    percent = round((total / 60) * 100, 1)
    cefr = get_cefr(int(total))

    # Get comment from lowest scoring LLM
    lowest_llm = min(done, key=lambda k: done[k].get("total", 0))
    comment = done[lowest_llm].get("summary", "")

    # Load or create workbook
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CEFR Results"
        # Write headers
        headers = ["Name"] + PARAMETERS + ["Total (/60)", "Percent (%)", "CEFR Level", "AI Comment"]
        ws.append(headers)

    # Write user row
    row = [username] + [avg_scores[p] for p in PARAMETERS] + [total, percent, cefr, comment]
    ws.append(row)
    wb.save(EXCEL_PATH)

# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route("/health", methods=["GET"])
def health():
    active = [name for name, key in API_KEYS.items() if key and not key.startswith("your_")]
    return jsonify({"status": "ok", "active_llms": active})

@app.route("/analyze", methods=["POST"])
def analyze():
    if "audio" not in request.files:
        return jsonify({"error": "No audio file provided"}), 400

    audio_file = request.files["audio"]

    try:
        # Step 1: Transcribe with Groq Whisper
        transcript, audio_duration = transcribe_audio(audio_file)

        # Step 2: Extract features
        features = extract_features(transcript, audio_duration)

        # Step 3: Build prompt
        prompt = build_prompt(transcript, features)

        # Step 4: Score with all available LLMs in parallel
        available = {
            name: key for name, key in API_KEYS.items()
            if key and not key.startswith("your_")
        }

        llm_results = {}
        threads = []
        for name, key in available.items():
            t = threading.Thread(target=score_with_llm,
                                 args=(name, prompt, key, llm_results))
            t.start()
            threads.append(t)
        for t in threads:
            t.join()

        save_to_excel(
            request.form.get("username", "Unknown"),
            llm_results,
            features
        )

        return jsonify({
            "transcript": transcript,
            "features": features,
            "llm_results": llm_results
        })

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/improve", methods=["POST"])
def improve():
    payload = request.get_json()
    transcript = payload.get("transcript", "")
    avg_scores = payload.get("avg_scores", {})

    weak_params = {p: s for p, s in avg_scores.items() if s <= 2}

    if not weak_params:
        return jsonify({
            "improvements": [],
            "overall_encouragement": "Great job! All your parameters scored above the improvement threshold. Keep practicing to reach the next CEFR level!"
        })

    prompt = build_improvement_prompt(transcript, weak_params, avg_scores)

    # Use Groq (fastest) for improvement tips
    groq_key = API_KEYS.get("Groq (Llama 3.3 70B)")
    if not groq_key:
        return jsonify({"error": "Groq API key not configured"}), 500

    try:
        raw = call_groq(prompt, groq_key)
        data = json.loads(raw)
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(debug=True, port=5000)